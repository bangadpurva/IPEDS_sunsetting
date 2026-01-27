import pandas as pd
from pathlib import Path
from typing import Dict, Union
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# CONFIG — EDIT FILE PATHS
# =========================
FILES_BY_YEAR: Dict[int, Union[str, Path]] = {
    2019: "data_stu/c2019_c.csv",   # <- update paths if needed
    2020: "data_stu/c2020_c.csv",
    2021: "data_stu/c2021_c.csv",
    2022: "data_stu/c2022_c.csv",
    2023: "data_stu/c2023_c.csv",
    2024: "data_stu/c2024_c.csv",
}

EXCEL_SHEET = None   # use first sheet; change if needed

COL_UNITID = "UNITID"
COL_CIP = "CIPCODE"
COL_AWLEVEL = "AWLEVEL"
COL_MAJORNUM = "MAJORNUM"
COL_TOTAL = "CTOTALT"

KEEP_PRIMARY_MAJOR_ONLY = True   # prevents double counting
PRIMARY_MAJOR_VALUE = 1

COMBINED_OUT = "data/completions_students_2019_2024.csv"
OUT_XLSX = "cip_awlevel_students_yoy_2019_2024.xlsx"
SHEET_NAME = "Student YoY Summary"

# =========================
# 1) BUILD COMBINED FILE
# =========================
frames = []

for year, path in FILES_BY_YEAR.items():
    path = Path(path)

    if path.suffix.lower() in [".xlsx", ".xls"]:
        df = pd.read_excel(path, sheet_name=EXCEL_SHEET)
    else:
        df = pd.read_csv(path)

    # Standardize columns
    for c in [COL_UNITID, COL_CIP, COL_AWLEVEL, COL_MAJORNUM]:
        if c in df.columns:
            df[c] = df[c].astype("string").str.strip()

    df[COL_TOTAL] = pd.to_numeric(df[COL_TOTAL], errors="coerce").fillna(0)
    df["YEAR"] = year

    keep_cols = [COL_UNITID, COL_CIP, COL_MAJORNUM, COL_AWLEVEL, COL_TOTAL, "YEAR"]
    df = df[keep_cols]

    frames.append(df)

combined = pd.concat(frames, ignore_index=True)

# keep first major only
if KEEP_PRIMARY_MAJOR_ONLY and COL_MAJORNUM in combined.columns:
    combined[COL_MAJORNUM] = pd.to_numeric(combined[COL_MAJORNUM], errors="coerce")
    combined = combined[combined[COL_MAJORNUM] == PRIMARY_MAJOR_VALUE]

Path("data").mkdir(exist_ok=True)
combined.to_csv(COMBINED_OUT, index=False)
print(f"Saved combined file → {COMBINED_OUT}")

# =========================
# 2) AGGREGATE: STUDENT COMPLETIONS
# =========================
agg = (
    combined.groupby([COL_CIP, COL_AWLEVEL, "YEAR"], as_index=False)
            .agg(total_students_completed=(COL_TOTAL, "sum"))
)

years = list(range(2019, 2025))

wide = (
    agg.pivot_table(index=[COL_CIP, COL_AWLEVEL], columns="YEAR",
                    values="total_students_completed", aggfunc="sum")
       .reindex(columns=years)
       .fillna(0)
       .reset_index()
)

# =========================
# 3) YOY COUNTS & YOY %
# =========================
for i in range(1, len(years)):
    y_prev, y_cur = years[i-1], years[i]
    wide[f"{y_cur}-{str(y_prev)[-2:]}"] = wide[y_cur] - wide[y_prev]
    wide[f"{y_cur}-{str(y_prev)[-2:]}%"] = (
        (wide[y_cur] - wide[y_prev]) / wide[y_prev].replace({0: pd.NA}) * 100
    )

pct_cols = [c for c in wide.columns if c.endswith("%")]
wide[pct_cols] = wide[pct_cols].round(2)

# =========================
# 4) FINAL OUTPUT STRUCTURE
# =========================
yoy_count_cols = [f"{years[i]}-{str(years[i-1])[-2:]}" for i in range(1, len(years))]
yoy_pct_cols = [c + "%" for c in yoy_count_cols]

out = wide[
    [COL_CIP, COL_AWLEVEL] +
    years +
    yoy_count_cols +
    yoy_pct_cols
]

# =========================
# 5) WRITE EXCEL WITH MERGED HEADERS
# =========================
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
    out.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=2)

wb = load_workbook(OUT_XLSX)
ws = wb[SHEET_NAME]

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
bold = Font(bold=True)

r_top, r_sub = 1, 2
total_cols = len(years)
yoy_cols = len(yoy_count_cols)

# Top headers
ws.cell(r_top, 1, "CIP Code")
ws.cell(r_top, 2, "AWLEVEL")

ws.cell(r_top, 3, "Total Students Completed")
ws.merge_cells(start_row=r_top, start_column=3, end_row=r_top, end_column=2 + total_cols)

ws.cell(r_top, 3 + total_cols, "Year over year Change count")
ws.merge_cells(start_row=r_top, start_column=3 + total_cols,
               end_row=r_top, end_column=2 + total_cols + yoy_cols)

ws.cell(r_top, 3 + total_cols + yoy_cols, "Year over year Change %")
ws.merge_cells(start_row=r_top, start_column=3 + total_cols + yoy_cols,
               end_row=r_top, end_column=2 + total_cols + 2*yoy_cols)

# Merge CIP & AWLEVEL vertically
ws.merge_cells(start_row=r_top, start_column=1, end_row=r_sub, end_column=1)
ws.merge_cells(start_row=r_top, start_column=2, end_row=r_sub, end_column=2)

# Subheaders
col = 3
for y in years:
    ws.cell(r_sub, col, str(y)); col += 1
for c in yoy_count_cols:
    ws.cell(r_sub, col, c); col += 1
for c in yoy_pct_cols:
    ws.cell(r_sub, col, c.replace("%","")); col += 1

# Format headers
for r in [r_top, r_sub]:
    for c in range(1, col):
        ws.cell(r, c).alignment = center
        ws.cell(r, c).font = bold

ws.freeze_panes = "C3"

for c in range(1, col):
    ws.column_dimensions[get_column_letter(c)].width = 12

wb.save(OUT_XLSX)
print(f"Saved student YoY Excel → {OUT_XLSX}")
