import pandas as pd
from pathlib import Path
from typing import Dict, Union
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
FILES_BY_YEAR: Dict[int, Union[str, Path]] = {
    2019: "data_uni/c2019_a.csv",
    2020: "data_uni/c2020_a.csv",
    2021: "data_uni/c2021_a.csv",
    2022: "data_uni/c2022_a.csv",
    2023: "data_uni/c2023_a.csv",
    2024: "data_uni/c2024_a.csv",
}

COL_UNITID = "UNITID"
COL_CIP = "CIPCODE"
COL_AWLEVEL = "AWLEVEL"
COL_MAJORNUM = "MAJORNUM"
COL_TOTAL = "CTOTALT"

KEEP_PRIMARY_MAJOR_ONLY = True
PRIMARY_MAJOR_VALUE = 1

COMBINED_OUT = "data_uni/completions_2019_2024.csv"
OUT_XLSX = "cip_awlevel_yoy_2019_2024.xlsx"
SHEET_NAME = "YoY Summary"

YEARS = list(range(2019, 2025))


def main():
    frames = []
    for year, path in FILES_BY_YEAR.items():
        df = pd.read_csv(path)

        for c in [COL_UNITID, COL_CIP, COL_AWLEVEL, COL_MAJORNUM]:
            if c in df.columns:
                df[c] = df[c].astype("string").str.strip()

        df[COL_TOTAL] = pd.to_numeric(df[COL_TOTAL], errors="coerce").fillna(0)
        df["YEAR"] = year

        keep_cols = [COL_UNITID, COL_CIP, COL_MAJORNUM, COL_AWLEVEL, COL_TOTAL, "YEAR"]
        frames.append(df[keep_cols])

    combined = pd.concat(frames, ignore_index=True)

    if KEEP_PRIMARY_MAJOR_ONLY and COL_MAJORNUM in combined.columns:
        combined[COL_MAJORNUM] = pd.to_numeric(combined[COL_MAJORNUM], errors="coerce")
        combined = combined[combined[COL_MAJORNUM] == PRIMARY_MAJOR_VALUE]

    Path("data_uni").mkdir(exist_ok=True)
    combined.to_csv(COMBINED_OUT, index=False)
    print(f"Saved combined file → {COMBINED_OUT}")

    agg = (
        combined.groupby([COL_CIP, COL_AWLEVEL, "YEAR"], as_index=False)
                .agg(total_completed=(COL_TOTAL, "sum"))
    )

    wide = (
        agg.pivot_table(index=[COL_CIP, COL_AWLEVEL], columns="YEAR",
                        values="total_completed", aggfunc="sum")
           .reindex(columns=YEARS)
           .fillna(0)
           .reset_index()
    )

    yoy_count_cols = []
    yoy_pct_cols = []
    for i in range(1, len(YEARS)):
        y_prev, y_cur = YEARS[i - 1], YEARS[i]
        diff_col = f"{y_cur}-{str(y_prev)[-2:]}"
        pct_col = f"{diff_col}%"
        wide[diff_col] = wide[y_cur] - wide[y_prev]
        wide[pct_col] = ((wide[y_cur] - wide[y_prev]) / wide[y_prev].replace({0: pd.NA})) * 100
        yoy_count_cols.append(diff_col)
        yoy_pct_cols.append(pct_col)

    wide[yoy_pct_cols] = wide[yoy_pct_cols].apply(pd.to_numeric, errors="coerce").round(2)

    out = wide[[COL_CIP, COL_AWLEVEL] + YEARS + yoy_count_cols + yoy_pct_cols]

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=2)

    wb = load_workbook(OUT_XLSX)
    ws = wb[SHEET_NAME]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    r_top, r_sub = 1, 2
    total_cols = len(YEARS)
    yoy_cols = len(yoy_count_cols)

    ws.cell(r_top, 1, "CIP Code")
    ws.cell(r_top, 2, "AWLEVEL")
    ws.cell(r_top, 3, "Total Completed")
    ws.merge_cells(start_row=r_top, start_column=3, end_row=r_top, end_column=2 + total_cols)
    ws.cell(r_top, 3 + total_cols, "Year over year Change count")
    ws.merge_cells(start_row=r_top, start_column=3 + total_cols,
                   end_row=r_top, end_column=2 + total_cols + yoy_cols)
    ws.cell(r_top, 3 + total_cols + yoy_cols, "Year over year Change %")
    ws.merge_cells(start_row=r_top, start_column=3 + total_cols + yoy_cols,
                   end_row=r_top, end_column=2 + total_cols + 2 * yoy_cols)

    ws.merge_cells(start_row=r_top, start_column=1, end_row=r_sub, end_column=1)
    ws.merge_cells(start_row=r_top, start_column=2, end_row=r_sub, end_column=2)

    col = 3
    for y in YEARS:
        ws.cell(r_sub, col, str(y)); col += 1
    for c in yoy_count_cols:
        ws.cell(r_sub, col, c); col += 1
    for c in yoy_pct_cols:
        ws.cell(r_sub, col, c.replace("%", "")); col += 1

    for r in [r_top, r_sub]:
        for c in range(1, col):
            ws.cell(r, c).alignment = center
            ws.cell(r, c).font = bold

    ws.freeze_panes = "C3"
    for c in range(1, col):
        ws.column_dimensions[get_column_letter(c)].width = 12

    wb.save(OUT_XLSX)
    print(f"Saved YoY Excel → {OUT_XLSX}")


if __name__ == "__main__":
    main()
