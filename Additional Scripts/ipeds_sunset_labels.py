"""
Sunsetting Programs (IPEDS 2019–2024)
====================================

Uses both IPEDS A-files (CIPCODE + AWLEVEL completions by institution) and
C-files (AWLEVELC total completions system-wide) to produce:

A) Program-level YoY (UNITID × CIPCODE × AWLEVEL)
B) National YoY (CIPCODE × AWLEVEL)
C) System-level YoY (AWLEVELC, from C-files)
D) Decline labels per UNITID × CIPCODE × AWLEVEL:
   "structural program exit", "credential-system decline",
   "field-specific decline", "stable/unclear"

Outputs written to the `out/` directory.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Union

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
FILES_A_BY_YEAR: Dict[int, Union[str, Path]] = {
    2019: "data_uni/c2019_a.csv",
    2020: "data_uni/c2020_a.csv",
    2021: "data_uni/c2021_a.csv",
    2022: "data_uni/c2022_a.csv",
    2023: "data_uni/c2023_a.csv",
    2024: "data_uni/c2024_a.csv",
}

FILES_C_BY_YEAR: Dict[int, Union[str, Path]] = {
    2019: "data_stu/c2019_c.csv",
    2020: "data_stu/c2020_c.csv",
    2021: "data_stu/c2021_c.csv",
    2022: "data_stu/c2022_c.csv",
    2023: "data_stu/c2023_c.csv",
    2024: "data_stu/c2024_c.csv",
}

OUT_DIR = Path("out")

COL_UNITID = "UNITID"
COL_CIP = "CIPCODE"
COL_AWLEVEL = "AWLEVEL"
COL_MAJORNUM = "MAJORNUM"
COL_TOTAL_A = "CTOTALT"

KEEP_PRIMARY_MAJOR_ONLY = True
PRIMARY_MAJOR_VALUE = 1

COL_AWLEVELC = "AWLEVELC"
COL_TOTAL_C = "CSTOTLT"

YEARS = list(range(2019, 2025))

STRUCTURAL_EXIT_MIN_ACTIVE_YEARS = 3
STRUCTURAL_EXIT_LAST2_ZERO = True
DECLINE_PCT_THRESHOLD = -10.0
DECLINE_ABS_THRESHOLD = -5
SYSTEM_DECLINE_THRESHOLD = -5.0

AWLEVEL_TO_AWLEVELC = {
    "01": "1", "1": "1",
    "02": "2", "2": "2",
    "03": "3", "3": "3",
    "05": "5", "5": "5",
    "07": "7", "7": "7",
    "09": "9", "9": "9",
    "04": "10", "06": "10", "08": "10", "10": "10",
    "11": "11", "12": "12",
    "17": "10", "18": "10", "19": "10", "20": "10", "21": "10",
}


# =========================
# HELPERS
# =========================

def _read_csv_clean(path: Union[str, Path]) -> pd.DataFrame:
    df = pd.read_csv(path, dtype="string", low_memory=False)
    for c in df.columns:
        if df[c].dtype == "string":
            df[c] = df[c].str.strip()
    return df


def _to_numeric_safe(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def add_yoy_columns(wide: pd.DataFrame, years: List[int]) -> pd.DataFrame:
    for i in range(1, len(years)):
        y_prev, y_cur = years[i - 1], years[i]
        wide[f"{y_cur}-{str(y_prev)[-2:]}"] = wide[y_cur] - wide[y_prev]
        wide[f"{y_cur}-{str(y_prev)[-2:]}%"] = (
            (wide[y_cur] - wide[y_prev]) / wide[y_prev].replace({0: pd.NA}) * 100
        )
    pct_cols = [c for c in wide.columns if isinstance(c, str) and c.endswith("%")]
    if pct_cols:
        wide[pct_cols] = wide[pct_cols].apply(pd.to_numeric, errors="coerce").round(2)
    return wide


def write_merged_header_excel(
    df: pd.DataFrame,
    out_path: Path,
    sheet_name: str,
    left_cols: List[str],
    years: List[int],
    yoy_count_cols: List[str],
    yoy_pct_cols: List[str],
    title_total: str = "Total Completed",
    title_yoy_count: str = "Year over year Change count",
    title_yoy_pct: str = "Year over year Change %",
):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=2)

    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    r_top, r_sub = 1, 2

    for i, col_name in enumerate(left_cols, start=1):
        ws.cell(r_top, i, col_name)
        ws.merge_cells(start_row=r_top, start_column=i, end_row=r_sub, end_column=i)

    start = len(left_cols) + 1
    total_cols = len(years)
    yoy_cols = len(yoy_count_cols)

    ws.cell(r_top, start, title_total)
    ws.merge_cells(start_row=r_top, start_column=start, end_row=r_top, end_column=start + total_cols - 1)

    ws.cell(r_top, start + total_cols, title_yoy_count)
    ws.merge_cells(
        start_row=r_top, start_column=start + total_cols,
        end_row=r_top, end_column=start + total_cols + yoy_cols - 1,
    )

    ws.cell(r_top, start + total_cols + yoy_cols, title_yoy_pct)
    ws.merge_cells(
        start_row=r_top, start_column=start + total_cols + yoy_cols,
        end_row=r_top, end_column=start + total_cols + 2 * yoy_cols - 1,
    )

    col = start
    for y in years:
        ws.cell(r_sub, col, str(y)); col += 1
    for c in yoy_count_cols:
        ws.cell(r_sub, col, c); col += 1
    for c in yoy_pct_cols:
        ws.cell(r_sub, col, c.replace("%", "")); col += 1

    for r in [r_top, r_sub]:
        for c in range(1, col):
            ws.cell(r, c).alignment = center
            ws.cell(r, c).font = bold

    ws.freeze_panes = f"{get_column_letter(len(left_cols)+1)}3"
    for c in range(1, col):
        ws.column_dimensions[get_column_letter(c)].width = 14 if c <= len(left_cols) else 12

    wb.save(out_path)


def classify_decline_row(
    row: pd.Series,
    years: List[int],
    system_yoy_lookup: Dict[str, Dict[str, float]],
) -> str:
    awlevelc = str(row.get("AWLEVELC_mapped", "")).strip()
    vals = np.array([row.get(y, 0) for y in years], dtype=float)
    active_years = int((vals > 0).sum())

    y_prev, y_cur = years[-2], years[-1]
    last_yoy = float(row.get(f"{y_cur}-{str(y_prev)[-2:]}", 0))
    last_yoy_pct = row.get(f"{y_cur}-{str(y_prev)[-2:]}%")
    try:
        last_yoy_pct = float(last_yoy_pct) if last_yoy_pct is not None and not pd.isna(last_yoy_pct) else np.nan
    except Exception:
        last_yoy_pct = np.nan

    if active_years >= STRUCTURAL_EXIT_MIN_ACTIVE_YEARS:
        if STRUCTURAL_EXIT_LAST2_ZERO and vals[-1] == 0 and vals[-2] == 0 and vals[:-2].max() > 0:
            return "structural program exit"
        if vals[-2] > 0 and vals[-1] == 0 and last_yoy <= -max(DECLINE_ABS_THRESHOLD, 1):
            return "structural program exit"

    meaningful_decline = (last_yoy <= DECLINE_ABS_THRESHOLD) or (
        not np.isnan(last_yoy_pct) and last_yoy_pct <= DECLINE_PCT_THRESHOLD
    )
    if not meaningful_decline:
        return "stable/unclear"

    sys = system_yoy_lookup.get(awlevelc, {})
    sys_last_yoy_pct = sys.get(f"{y_cur}-{str(y_prev)[-2:]}%")
    if sys_last_yoy_pct is not None and not pd.isna(sys_last_yoy_pct) and sys_last_yoy_pct <= SYSTEM_DECLINE_THRESHOLD:
        return "credential-system decline"

    return "field-specific decline"


# =========================
# MAIN
# =========================

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    yoy_count_cols = [f"{YEARS[i]}-{str(YEARS[i-1])[-2:]}" for i in range(1, len(YEARS))]
    yoy_pct_cols = [c + "%" for c in yoy_count_cols]

    # --- Load A files ---
    frames_a = []
    for year, path in FILES_A_BY_YEAR.items():
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Missing A file for {year}: {path}")
        df = _read_csv_clean(path)
        needed = [COL_UNITID, COL_CIP, COL_AWLEVEL, COL_TOTAL_A]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            raise ValueError(f"{path} is missing columns: {missing}")
        df[COL_UNITID] = df[COL_UNITID].astype("string").str.strip()
        df[COL_CIP] = df[COL_CIP].astype("string").str.strip()
        df[COL_AWLEVEL] = df[COL_AWLEVEL].astype("string").str.strip()
        if COL_MAJORNUM in df.columns:
            df[COL_MAJORNUM] = df[COL_MAJORNUM].astype("string").str.strip()
        df[COL_TOTAL_A] = _to_numeric_safe(df[COL_TOTAL_A]).fillna(0)
        df["YEAR"] = year
        keep_cols = [COL_UNITID, COL_CIP, COL_AWLEVEL, COL_TOTAL_A, "YEAR"]
        if COL_MAJORNUM in df.columns:
            keep_cols.insert(2, COL_MAJORNUM)
        frames_a.append(df[keep_cols])

    a_combined = pd.concat(frames_a, ignore_index=True)
    if KEEP_PRIMARY_MAJOR_ONLY and COL_MAJORNUM in a_combined.columns:
        a_combined[COL_MAJORNUM] = _to_numeric_safe(a_combined[COL_MAJORNUM])
        a_combined = a_combined[a_combined[COL_MAJORNUM] == PRIMARY_MAJOR_VALUE]

    a_combined["AWLEVELC_mapped"] = (
        a_combined[COL_AWLEVEL].astype("string").str.strip().map(AWLEVEL_TO_AWLEVELC)
    )
    a_combined.to_csv(OUT_DIR / "combined_A_completions_2019_2024.csv", index=False)

    # --- Load C files ---
    frames_c = []
    for year, path in FILES_C_BY_YEAR.items():
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"Missing C file for {year}: {path}")
        df = _read_csv_clean(path)
        needed = [COL_UNITID, COL_AWLEVELC, COL_TOTAL_C]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            raise ValueError(f"{path} is missing columns: {missing}")
        df[COL_UNITID] = df[COL_UNITID].astype("string").str.strip()
        df[COL_AWLEVELC] = df[COL_AWLEVELC].astype("string").str.strip()
        df[COL_TOTAL_C] = _to_numeric_safe(df[COL_TOTAL_C]).fillna(0)
        df["YEAR"] = year
        frames_c.append(df[[COL_UNITID, COL_AWLEVELC, COL_TOTAL_C, "YEAR"]])

    c_combined = pd.concat(frames_c, ignore_index=True)
    c_combined.to_csv(OUT_DIR / "combined_C_system_totals_2019_2024.csv", index=False)

    # --- Output A: Program YoY by UNITID × CIP × AWLEVEL ---
    agg_uni = (
        a_combined.groupby([COL_UNITID, COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped", "YEAR"], as_index=False)
        .agg(total_completed=(COL_TOTAL_A, "sum"))
    )
    wide_uni = (
        agg_uni.pivot_table(
            index=[COL_UNITID, COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped"],
            columns="YEAR", values="total_completed", aggfunc="sum",
        )
        .reindex(columns=YEARS).fillna(0).reset_index()
    )
    wide_uni = add_yoy_columns(wide_uni, YEARS)
    cols_uni = [COL_UNITID, COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped"] + YEARS + yoy_count_cols + yoy_pct_cols
    out_uni = wide_uni[cols_uni]

    out_path_uni = OUT_DIR / "program_yoy_by_uni_cip_awlevel_2019_2024.xlsx"
    write_merged_header_excel(
        df=out_uni, out_path=out_path_uni, sheet_name="Program YoY (UNITID)",
        left_cols=["UNITID", "CIP Code", "AWLEVEL", "AWLEVELC_mapped"],
        years=YEARS, yoy_count_cols=yoy_count_cols, yoy_pct_cols=yoy_pct_cols,
        title_total="Total Completed (A files)",
    )

    # --- Output B: National YoY by CIP × AWLEVEL ---
    agg_nat = (
        a_combined.groupby([COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped", "YEAR"], as_index=False)
        .agg(total_completed=(COL_TOTAL_A, "sum"))
    )
    wide_nat = (
        agg_nat.pivot_table(
            index=[COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped"],
            columns="YEAR", values="total_completed", aggfunc="sum",
        )
        .reindex(columns=YEARS).fillna(0).reset_index()
    )
    wide_nat = add_yoy_columns(wide_nat, YEARS)
    cols_nat = [COL_CIP, COL_AWLEVEL, "AWLEVELC_mapped"] + YEARS + yoy_count_cols + yoy_pct_cols
    out_nat = wide_nat[cols_nat]

    out_path_nat = OUT_DIR / "national_yoy_by_cip_awlevel_2019_2024.xlsx"
    write_merged_header_excel(
        df=out_nat, out_path=out_path_nat, sheet_name="National YoY (CIP×AWLEVEL)",
        left_cols=["CIP Code", "AWLEVEL", "AWLEVELC_mapped"],
        years=YEARS, yoy_count_cols=yoy_count_cols, yoy_pct_cols=yoy_pct_cols,
        title_total="Total Completed (National, A files)",
    )

    # --- Output C: System YoY by AWLEVELC ---
    agg_sys = (
        c_combined.groupby([COL_AWLEVELC, "YEAR"], as_index=False)
        .agg(system_total_completed=(COL_TOTAL_C, "sum"))
    )
    wide_sys = (
        agg_sys.pivot_table(
            index=[COL_AWLEVELC], columns="YEAR",
            values="system_total_completed", aggfunc="sum",
        )
        .reindex(columns=YEARS).fillna(0).reset_index()
    )
    wide_sys = add_yoy_columns(wide_sys, YEARS)
    cols_sys = [COL_AWLEVELC] + YEARS + yoy_count_cols + yoy_pct_cols
    out_sys = wide_sys[cols_sys]

    out_path_sys = OUT_DIR / "system_yoy_by_awlevelc_2019_2024.xlsx"
    write_merged_header_excel(
        df=out_sys, out_path=out_path_sys, sheet_name="System YoY (AWLEVELC)",
        left_cols=["AWLEVELC"], years=YEARS,
        yoy_count_cols=yoy_count_cols, yoy_pct_cols=yoy_pct_cols,
        title_total="System Total Completed (C files)",
    )

    # Build system-level YoY lookup for decline labeling
    system_yoy_lookup: Dict[str, Dict[str, float]] = {}
    for _, r in wide_sys.iterrows():
        awc = str(r[COL_AWLEVELC]).strip()
        system_yoy_lookup[awc] = {}
        for c in yoy_pct_cols:
            v = r.get(c)
            try:
                system_yoy_lookup[awc][c] = float(v) if v is not None and not pd.isna(v) else np.nan
            except Exception:
                system_yoy_lookup[awc][c] = np.nan

    # --- Output D: Sunsetting Labels by UNITID × CIP × AWLEVEL ---
    labels_df = out_uni.copy()
    labels_df["decline_label"] = labels_df.apply(
        lambda r: classify_decline_row(r, YEARS, system_yoy_lookup), axis=1
    )

    last_pair = f"{YEARS[-1]}-{str(YEARS[-2])[-2:]}"
    labels_df["last_yoy_change"] = labels_df[last_pair]
    labels_df["last_yoy_pct"] = labels_df[f"{last_pair}%"]

    out_path_labels = OUT_DIR / "sunsetting_labels_by_uni_cip_awlevel_2019_2024.xlsx"
    with pd.ExcelWriter(out_path_labels, engine="openpyxl") as writer:
        labels_df.to_excel(writer, sheet_name="Sunsetting Labels", index=False)

    print("DONE.")
    print(f"- {out_path_uni}")
    print(f"- {out_path_nat}")
    print(f"- {out_path_sys}")
    print(f"- {out_path_labels}")


if __name__ == "__main__":
    main()
