import pandas as pd
from pathlib import Path
from typing import Dict, Union
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# =========================
# CONFIG
# =========================
FILES_BY_YEAR: Dict[int, Union[str, Path]] = {
    2019: "data_stu/c2019_c.csv",
    2020: "data_stu/c2020_c.csv",
    2021: "data_stu/c2021_c.csv",
    2022: "data_stu/c2022_c.csv",
    2023: "data_stu/c2023_c.csv",
    2024: "data_stu/c2024_c.csv",
}

COL_UNITID = "UNITID"
COL_AWLEVELC = "AWLEVELC"
COL_TOTAL_COMPLETIONS = "CSTOTLT"

COMBINED_OUT = "data_stu/completions_students_2019_2024.csv"
OUT_XLSX = "awlevelc_students_yoy_2019_2024.xlsx"
SHEET_NAME = "Student YoY Summary"

YEARS = list(range(2019, 2025))


def main():
    frames = []
    for year, path in FILES_BY_YEAR.items():
        path = Path(path)
        df = pd.read_csv(path, low_memory=False)

        if COL_UNITID in df.columns:
            df[COL_UNITID] = df[COL_UNITID].astype("string").str.strip()
        if COL_AWLEVELC in df.columns:
            df[COL_AWLEVELC] = df[COL_AWLEVELC].astype("string").str.strip()

        df[COL_TOTAL_COMPLETIONS] = pd.to_numeric(df[COL_TOTAL_COMPLETIONS], errors="coerce").fillna(0)
        df["YEAR"] = year

        keep_cols = [COL_UNITID, COL_AWLEVELC, COL_TOTAL_COMPLETIONS, "YEAR"]
        missing = [c for c in keep_cols if c not in df.columns]
        if missing:
            raise ValueError(f"{path} is missing expected columns: {missing}")

        frames.append(df[keep_cols])

    Path("data_stu").mkdir(exist_ok=True)
    combined = pd.concat(frames, ignore_index=True)
    combined.to_csv(COMBINED_OUT, index=False)
    print(f"Saved combined file → {COMBINED_OUT}")

    agg = (
        combined.groupby([COL_AWLEVELC, "YEAR"], as_index=False)
                .agg(total_completed=(COL_TOTAL_COMPLETIONS, "sum"))
    )

    wide = (
        agg.pivot_table(index=[COL_AWLEVELC], columns="YEAR", values="total_completed", aggfunc="sum")
           .reindex(columns=YEARS)
           .fillna(0)
           .reset_index()
    )

    yoy_count_cols = []
    yoy_pct_cols = []
    for i in range(1, len(YEARS)):
        y_prev, y_cur = YEARS[i - 1], YEARS[i]
        diff_col = f"{str(y_cur)[-2:]}-{str(y_prev)[-2:]}"
        pct_col = f"{diff_col}%"
        wide[diff_col] = wide[y_cur] - wide[y_prev]
        wide[pct_col] = (wide[diff_col] / wide[y_prev].replace({0: pd.NA})) * 100
        yoy_count_cols.append(diff_col)
        yoy_pct_cols.append(pct_col)

    wide[yoy_pct_cols] = wide[yoy_pct_cols].round(2)

    out = wide[[COL_AWLEVELC] + YEARS + yoy_count_cols + yoy_pct_cols]

    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name=SHEET_NAME, index=False, header=False, startrow=2)

    wb = load_workbook(OUT_XLSX)
    ws = wb[SHEET_NAME]

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    r_top, r_sub = 1, 2
    total_cols = len(YEARS)
    yoy_cols = len(yoy_count_cols)

    ws.cell(r_top, 1, "AWLEVELC")
    ws.cell(r_top, 2, "Total Completed")
    ws.merge_cells(start_row=r_top, start_column=2, end_row=r_top, end_column=1 + total_cols)
    ws.cell(r_top, 2 + total_cols, "Year over year Change count")
    ws.merge_cells(start_row=r_top, start_column=2 + total_cols,
                   end_row=r_top, end_column=1 + total_cols + yoy_cols)
    ws.cell(r_top, 2 + total_cols + yoy_cols, "Year over year Change %")
    ws.merge_cells(start_row=r_top, start_column=2 + total_cols + yoy_cols,
                   end_row=r_top, end_column=1 + total_cols + 2 * yoy_cols)

    ws.merge_cells(start_row=r_top, start_column=1, end_row=r_sub, end_column=1)

    col = 2
    for y in YEARS:
        ws.cell(r_sub, col, str(y)); col += 1
    for c in yoy_count_cols:
        ws.cell(r_sub, col, c); col += 1
    for c in yoy_pct_cols:
        ws.cell(r_sub, col, c.replace("%", "")); col += 1

    for r in [r_top, r_sub]:
        for c in range(1, col):
            ws.cell(r, c).alignment = center
            ws.cell(r, c).font = bold

    ws.freeze_panes = "B3"
    for c in range(1, col):
        ws.column_dimensions[get_column_letter(c)].width = 14

    wb.save(OUT_XLSX)
    print(f"Saved student YoY Excel → {OUT_XLSX}")


if __name__ == "__main__":
    main()
